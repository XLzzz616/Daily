#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File         :bdqx_gmsj_v1.py
@Time         :2020/04/23 11:14:46
@Auther       :Xuz
@Version      :1.1
@Notes        :下载百度迁徙-迁徙规模指数数据
              v1.0 迁徙规模指数：反映迁入或迁出人口规模，城市间可横向对比
              v1.1 xlwt长度限制,改用openpyxl
'''
from urllib import request
import re
import openpyxl
import requests
import time

#city information
ID=[341700,341200,340100,340600,340400,341000,340500,340800,340700,340300,341100,341500,340200,341300,341800,341600,820000,110000,350900,350300,350100,350500,350400,350800,350700,350200,350600,622900,620500,621100,621000,623000,620900,620700,620300,620600,620100,620400,620800,621200,620200,445100,441600,445200,440900,441400,441800,441500,440200,441700,445300,440800,440600,440100,440700,441200,442000,440400,441300,441900,440300,440500,451100,450100,450200,450300,451000,450500,450600,450800,451200,450700,450400,450900,451400,451300,520400,520500,520200,522600,522700,522300,520600,520300,520100,460100,460300,460200,469006,469007,469024,469027,469001,469029,460400,469021,469022,469005,469026,469023,469030,469002,469025,469028,130600,130900,130800,130400,131100,131000,130300,130100,130200,130500,130700,410600,410800,410200,411100,411300,410400,411200,411400,419001,410700,411500,410100,410500,410300,411700,411600,411000,410900,232700,230200,230300,231100,230400,230900,230500,230700,230800,231200,231000,230100,230600,429004,420300,420600,429021,421200,422800,420900,429006,420800,420500,420100,429005,421000,420700,421300,421100,420200,430200,430300,430100,430600,433100,430700,430400,431300,431200,431000,430900,431100,430800,430500,222400,220500,220700,220300,220400,220800,220200,220100,220600,320300,320200,320500,320400,320800,321000,320900,320700,321300,320100,321100,321200,320600,360200,360100,361000,360500,360300,360400,361100,360800,360900,360700,360600,210500,211300,210600,210400,210900,211400,210700,211100,211200,210800,210300,210200,211000,210100,150700,152900,150800,150200,150500,152500,150100,150400,150600,152200,150900,150300,640400,640200,640300,640500,640100,632200,630200,632500,632800,632300,630100,632600,632700,371100,370200,370100,370700,371300,370800,370500,370600,371700,371600,371400,370300,370400,370900,371500,371000,140300,140500,141000,140200,140400,141100,140100,140700,140600,140800,140900,610600,610800,610100,610200,610900,611000,610400,610300,610700,610500,310000,510800,511900,511700,510700,511500,511400,513400,511100,510300,510100,511800,510600,513200,513300,511300,511600,510500,511000,510400,510900,512000,719003,710600,710100,710400,710300,710500,719001,710200,719002,719010,719008,719011,719009,719013,719007,719012,719005,719006,719004,719014,120000,540100,540300,540400,540500,540200,540600,542500,810000,652800,652700,652300,650500,653200,653100,654200,654300,654000,659009,652900,659002,653000,659005,659006,659008,650100,650400,650200,659004,659001,659003,659007,532900,532300,532600,530800,530400,532500,533400,530700,533300,530500,533100,530300,530900,532800,530600,530100,330700,331000,330400,331100,330800,330200,330600,330100,330500,330300,330900,500000]
name=['池州市','阜阳市','合肥市','淮北市','淮南市','黄山市','马鞍山市','安庆市','铜陵市','蚌埠市','滁州市','六安市','芜湖市','宿州市','宣城市','亳州市','澳门','北京市','宁德市','莆田市','福州市','泉州市','三明市','龙岩市','南平市','厦门市','漳州市','临夏回族自治州','天水市','定西市','庆阳市','甘南藏族自治州','酒泉市','张掖市','金昌市','武威市','兰州市','白银市','平凉市','陇南市','嘉峪关市','潮州市','河源市','揭阳市','茂名市','梅州市','清远市','汕尾市','韶关市','阳江市','云浮市','湛江市','佛山市','广州市','江门市','肇庆市','中山市','珠海市','惠州市','东莞市','深圳市','汕头市','贺州市','南宁市','柳州市','桂林市','百色市','北海市','防城港市','贵港市','河池市','钦州市','梧州市','玉林市','崇左市','来宾市','安顺市','毕节市','六盘水市','黔东南苗族侗族自治州','黔南布依族苗族自治州','黔西南布依族苗族自治州','铜仁市','遵义市','贵阳市','海口市','三沙市','三亚市','万宁市','东方市','临高县','乐东黎族自治县','五指山市','保亭黎族苗族自治县','儋州市','定安县','屯昌县','文昌市','昌江黎族自治县','澄迈县','琼中黎族苗族自治县','琼海市','白沙黎族自治县','陵水黎族自治县','保定市','沧州市','承德市','邯郸市','衡水市','廊坊市','秦皇岛市','石家庄市','唐山市','邢台市','张家口市','鹤壁市','焦作市','开封市','漯河市','南阳市','平顶山市','三门峡市','商丘市','济源市','新乡市','信阳市','郑州市','安阳市','洛阳市','驻马店市','周口市','许昌市','濮阳市','大兴安岭地区','齐齐哈尔市','鸡西市','黑河市','鹤岗市','七台河市','双鸭山市','伊春市','佳木斯市','绥化市','牡丹江市','哈尔滨市','大庆市','仙桃市','十堰市','襄阳市','神农架林区','咸宁市','恩施土家族苗族自治州','孝感市','天门市','荆门市','宜昌市','武汉市','潜江市','荆州市','鄂州市','随州市','黄冈市','黄石市','株洲市','湘潭市','长沙市','岳阳市','湘西土家族苗族自治州','常德市','衡阳市','娄底市','怀化市','郴州市','益阳市','永州市','张家界市','邵阳市','延边朝鲜族自治州','通化市','松原市','四平市','辽源市','白城市','吉林市','长春市','白山市','徐州市','无锡市','苏州市','常州市','淮安市','扬州市','盐城市','连云港市','宿迁市','南京市','镇江市','泰州市','南通市','景德镇市','南昌市','抚州市','新余市','萍乡市','九江市','上饶市','吉安市','宜春市','赣州市','鹰潭市','本溪市','朝阳市','丹东市','抚顺市','阜新市','葫芦岛市','锦州市','盘锦市','铁岭市','营口市','鞍山市','大连市','辽阳市','沈阳市','呼伦贝尔市','阿拉善盟','巴彦淖尔市','包头市','通辽市','锡林郭勒盟','呼和浩特市','赤峰市','鄂尔多斯市','兴安盟','乌兰察布市','乌海市','固原市','石嘴山市','吴忠市','中卫市','银川市','海北藏族自治州','海东市','海南藏族自治州','海西蒙古族藏族自治州','黄南藏族自治州','西宁市','果洛藏族自治州','玉树藏族自治州','日照市','青岛市','济南市','潍坊市','临沂市','济宁市','东营市','烟台市','菏泽市','滨州市','德州市','淄博市','枣庄市','泰安市','聊城市','威海市','阳泉市','晋城市','临汾市','大同市','长治市','吕梁市','太原市','晋中市','朔州市','运城市','忻州市','延安市','榆林市','西安市','铜川市','安康市','商洛市','咸阳市','宝鸡市','汉中市','渭南市','上海市','广元市','巴中市','达州市','绵阳市','宜宾市','眉山市','凉山彝族自治州','乐山市','自贡市','成都市','雅安市','德阳市','阿坝藏族羌族自治州','甘孜藏族自治州','南充市','广安市','泸州市','内江市','攀枝花市','遂宁市','资阳市','嘉义市','桃园市','台北市','台中市','新北市','台南市','基隆市','高雄市','新竹市','嘉义县','云林县','屏东县','南投县','花莲县','彰化县','台东县','宜兰县','苗栗县','新竹县','澎湖县','天津市','拉萨市','昌都市','林芝市','山南市','日喀则市','那曲市','阿里地区','香港','巴音郭楞蒙古自治州','博尔塔拉蒙古自治州','昌吉回族自治州','哈密市','和田地区','喀什地区','塔城地区','阿勒泰地区','伊犁哈萨克自治州','昆玉市','阿克苏地区','阿拉尔市','克孜勒苏柯尔克孜自治州','北屯市','铁门关市','可克达拉市','乌鲁木齐市','吐鲁番市','克拉玛依市','五家渠市','石河子市','图木舒克市','双河市','大理白族自治州','楚雄彝族自治州','文山壮族苗族自治州','普洱市','玉溪市','红河哈尼族彝族自治州','迪庆藏族自治州','丽江市','怒江傈僳族自治州','保山市','德宏傣族景颇族自治州','曲靖市','临沧市','西双版纳傣族自治州','昭通市','昆明市','金华市','台州市','嘉兴市','丽水市','衢州市','宁波市','绍兴市','杭州市','湖州市','温州市','舟山市','重庆市']

#request parameters
headers = {"User-agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0"}
opener = request.build_opener()
opener.add_headers = [headers]
request.install_opener(opener)

def getData(tarUrl,patList):    #根据url和数据格式获取数据，返回结果list
    connectUrl=tarUrl
    data=requests.get(connectUrl)
    res=[]
    for i in range(0,len(patList)):
        result = re.compile(patList[i]).findall(str(data.content.decode("utf-8")).encode("utf-8").decode("unicode_escape"))      
        res.append(result)
    return res

def scaleData(pFilename,startDate,endDate): #存储路径，起止时间
    f = openpyxl.Workbook()
    sheet1 = f.active
    sheet1.title="2019年迁徙规模指数"
    #sheet1=f.create_sheet(title="2019年迁徙规模指数")    
    sheet2=f.create_sheet(title="2020年迁徙规模指数")

    sDate=startDate
    eDate=endDate
    filename=pFilename
    
    #2019和2020年数据格式
    patDate2019 = '"(2019.*?)":\d+\.*\d.*?\D'
    patData2019 = '"2019.*?":(\d+\.*\d.*?)\D'
    patDate2020 = '"(2020.*?)":\d+\.*\d.*?\D'
    patData2020 = '"2020.*?":(\d+\.*\d.*?)\D'  
    #规模指数总pat
    patScale=[patDate2019,patData2019,patDate2020,patData2020]

    #全国数据
    countryUrlin="http://huiyan.baidu.com/migration/historycurve.jsonp?dt=country&id=0&type=move_in&startDate="+str(sDate)+"&endDate="+str(eDate)+"&callback=jsonp"
    resCouList1=getData(countryUrlin,patScale)
    row0_19=resCouList1[0]
    row1_19=resCouList1[1]
    row0_20=resCouList1[2]
    row1_20=resCouList1[3]
    sheet1.cell(1,1,u'城市')
    sheet1.cell(2,1,u'全国')
    sheet2.cell(1,1,u'城市')
    sheet2.cell(2,1,u'全国')
    for i1 in range(0,len(row0_19)):
        sheet1.cell(1,2*i1+2,'t'+row0_19[i1]+'迁入') 
        sheet1.cell(2,2*i1+2,float(row1_19[i1])*10000)
    for i2 in range(0,len(row0_20)):
        sheet2.cell(1,2*i2+2,'t'+row0_20[i2]+'迁入') 
        sheet2.cell(2,2*i2+2,float(row1_20[i2])*10000)
    countryUrlout="http://huiyan.baidu.com/migration/historycurve.jsonp?dt=country&id=0&type=move_out&startDate="+str(sDate)+"&endDate="+str(eDate)+"&callback=jsonp"  
    resCouList2=getData(countryUrlout,patScale)
    r2ow0_19=resCouList2[0]
    r2ow1_19=resCouList2[1]
    r2ow0_20=resCouList2[2]
    r2ow1_20=resCouList2[3]
    for i1 in range(0,len(r2ow0_19)):
        sheet1.cell(1,2*i1+3,'t'+r2ow0_19[i1]+'迁出') 
        sheet1.cell(2,2*i1+3,float(r2ow1_19[i1])*10000)
    for i2 in range(0,len(r2ow0_20)):
        sheet2.cell(1,2*i2+3,'t'+r2ow0_20[i2]+'迁出') 
        sheet2.cell(2,2*i2+3,float(r2ow1_20[i2])*10000)
    
    #迁入数据
    for i in range(0,len(ID)):
        cityUrlin = "http://huiyan.baidu.com/migration/historycurve.jsonp?dt=city&id="+str(ID[i])+"&type=move_in&startDate="+str(sDate)+"&endDate="+str(eDate)+"&callback=jsonp"
        reslistIn=getData(cityUrlin,patScale)
        row19_0=reslistIn[0]
        row19_i=reslistIn[1]
        row20_0=reslistIn[2]
        row20_i=reslistIn[3]
        #写入城市
        sheet1.cell(i+3,1,name[i]) 
        sheet2.cell(i+3,1,name[i])
        for i1 in range(0,len(row19_0)):
            sheet1.cell(i+3,2*i1+2,float(row19_i[i1])*10000)  
        for i2 in range(0,len(row20_0)):
            sheet2.cell(i+3,2*i2+2,float(row20_i[i2])*10000)
        f.save(filename)
        print (name[i],"in done")
        #time.sleep(1)
        #迁出数据    
        cityUrlout = "http://huiyan.baidu.com/migration/historycurve.jsonp?dt=city&id="+str(ID[i])+"&type=move_out&startDate="+str(sDate)+"&endDate="+str(eDate)+"&callback=jsonp"
        reslistOut=getData(cityUrlout,patScale)
        r2ow19_0=reslistOut[0]
        r2ow19_i=reslistOut[1]
        r2ow20_0=reslistOut[2]
        r2ow20_i=reslistOut[3]
        for i1 in range(0,len(r2ow19_0)):
            sheet1.cell(i+3,2*i1+3,float(r2ow19_i[i1])*10000)
        for i2 in range(0,len(r2ow20_0)):
            sheet2.cell(i+3,2*i2+3,float(r2ow20_i[i2])*10000)
        f.save(filename)
        cleft=len(ID)-i 
        print (name[i],"out done. 剩余",cleft,'个')
        time.sleep(1)
    
    print("规模数据抓取成功")

if __name__=='__main__':    
    startdate='20190112'
    enddate='20200506'
    filename = 'F:/DataGet/BDqianxi/'+'scaledata_'+str(enddate)+'.xlsx'    
    print('开始抓取')
    scaleData(filename,startdate,enddate)
    print('结束抓取')
    #time.sleep(300)